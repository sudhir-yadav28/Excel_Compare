"""
Excel File Comparator - Streamlit App
Compares two Excel files cell-by-cell and highlights matches (green) vs mismatches (yellow).
"""

import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ── Color constants ────────────────────────────────────────────────────────────
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


# ── Core comparison logic ──────────────────────────────────────────────────────

def align_dataframes(df_correct: pd.DataFrame, df_incorrect: pd.DataFrame, key_col: str | None):
    """
    Align both DataFrames on a common set of columns (and optionally a key column).
    Returns (correct_aligned, incorrect_aligned, common_cols, warnings).
    """
    warnings = []

    # Find common columns
    common_cols = [c for c in df_correct.columns if c in df_incorrect.columns]
    extra_correct   = [c for c in df_correct.columns   if c not in df_incorrect.columns]
    extra_incorrect = [c for c in df_incorrect.columns if c not in df_correct.columns]

    if extra_correct:
        warnings.append(f"Columns only in Correct File (ignored): {extra_correct}")
    if extra_incorrect:
        warnings.append(f"Columns only in Incorrect File (ignored): {extra_incorrect}")

    if not common_cols:
        raise ValueError("No common columns found between the two files.")

    if key_col:
        if key_col not in df_correct.columns:
            raise ValueError(f"Key column '{key_col}' not found in Correct File.")
        if key_col not in df_incorrect.columns:
            raise ValueError(f"Key column '{key_col}' not found in Incorrect File.")

        df_correct   = df_correct.set_index(key_col)
        df_incorrect = df_incorrect.set_index(key_col)

        # Align on the shared index values
        shared_index = df_correct.index.intersection(df_incorrect.index)
        only_correct   = df_correct.index.difference(df_incorrect.index).tolist()
        only_incorrect = df_incorrect.index.difference(df_correct.index).tolist()

        if only_correct:
            warnings.append(f"Rows in Correct File only (key '{key_col}'): {only_correct[:10]}"
                            + ("…" if len(only_correct) > 10 else ""))
        if only_incorrect:
            warnings.append(f"Rows in Incorrect File only (key '{key_col}'): {only_incorrect[:10]}"
                            + ("…" if len(only_incorrect) > 10 else ""))

        data_cols = [c for c in common_cols if c != key_col]
        correct_aligned   = df_correct.loc[shared_index, data_cols]
        incorrect_aligned = df_incorrect.loc[shared_index, data_cols]
    else:
        # Align on positional index
        min_rows = min(len(df_correct), len(df_incorrect))
        if len(df_correct) != len(df_incorrect):
            warnings.append(
                f"Row count differs: Correct={len(df_correct)}, Incorrect={len(df_incorrect)}. "
                f"Comparing first {min_rows} rows."
            )
        correct_aligned   = df_correct.iloc[:min_rows][common_cols].reset_index(drop=True)
        incorrect_aligned = df_incorrect.iloc[:min_rows][common_cols].reset_index(drop=True)

    return correct_aligned, incorrect_aligned, warnings


def compare_dataframes(correct: pd.DataFrame, incorrect: pd.DataFrame, case_sensitive: bool = True):
    """
    Build a boolean mask: True = values match, False = mismatch.
    Treats NaN == NaN as a match.
    """
    c = correct.fillna("__NaN__").astype(str)
    i = incorrect.fillna("__NaN__").astype(str)
    if not case_sensitive:
        c = c.apply(lambda col: col.str.lower())
        i = i.apply(lambda col: col.str.lower())
    match_mask = c == i
    return match_mask


def build_output_excel(incorrect: pd.DataFrame, match_mask: pd.DataFrame) -> bytes:
    """
    Write the incorrect DataFrame to an in-memory xlsx, then colour each cell:
      - GREEN  → values match
      - YELLOW → values differ
      - GRAY   → cell is outside the comparison area (shouldn't happen, but safe)
    Returns raw bytes of the xlsx file.
    """
    output = io.BytesIO()

    # Write with openpyxl engine so we can manipulate cells afterwards
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        incorrect.to_excel(writer, index=False, sheet_name="Comparison")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Row 1 is the header; data starts at row 2
    # No index column, so data starts at col 1
    for row_idx, row_label in enumerate(incorrect.index):
        for col_idx, col_label in enumerate(incorrect.columns):
            excel_row = row_idx + 2          # +1 for header, +1 for 1-based
            excel_col = col_idx + 1          # 1-based, no index column
            cell = ws.cell(row=excel_row, column=excel_col)

            try:
                is_match = match_mask.loc[row_label, col_label]
                cell.fill = GREEN if is_match else YELLOW
            except KeyError:
                cell.fill = GRAY

    result = io.BytesIO()
    wb.save(result)
    result.seek(0)
    return result.read()


def get_diff_preview(correct: pd.DataFrame, incorrect: pd.DataFrame, match_mask: pd.DataFrame,
                     max_rows: int = 100) -> pd.DataFrame:
    """
    Return a DataFrame of differing cells for display in the UI.
    Columns: Row, Column, Correct Value, Incorrect Value
    """
    rows = []
    for row_label in incorrect.index:
        for col_label in incorrect.columns:
            try:
                if not match_mask.loc[row_label, col_label]:
                    rows.append({
                        "Row":             row_label,
                        "Column":          col_label,
                        "Correct Value":   correct.loc[row_label, col_label],
                        "Incorrect Value": incorrect.loc[row_label, col_label],
                    })
            except KeyError:
                pass
            if len(rows) >= max_rows:
                return pd.DataFrame(rows)
    return pd.DataFrame(rows)


# ── Streamlit UI ───────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Excel Comparator", page_icon="📊", layout="wide")
    st.title("📊 Excel File Comparator")
    st.markdown(
        "Upload two Excel files to compare them cell-by-cell. "
        "**Green** = match · **Yellow** = mismatch."
    )

    # ── File upload section ────────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        correct_file = st.file_uploader("✅ Correct File (.xlsx)", type=["xlsx"])
    with col2:
        incorrect_file = st.file_uploader("❌ Incorrect File (.xlsx)", type=["xlsx"])

    # ── Options ────────────────────────────────────────────────────────────────
    with st.expander("⚙️ Options"):
        key_col_input = st.text_input(
            "Key column for row matching (leave blank to use row index)",
            value="",
            help="Enter a column name whose values uniquely identify each row (e.g. 'ID', 'OrderNo')."
        )
        key_col = key_col_input.strip() if key_col_input.strip() else None
        case_sensitive = st.checkbox("Case-sensitive comparison", value=True,
                                     help="When checked, 'Apple' and 'apple' are treated as different values.")
        show_preview = st.checkbox("Show mismatch preview table", value=True)

    # ── Main action ────────────────────────────────────────────────────────────
    if st.button("🔍 Compare & Generate Report", type="primary",
                 disabled=(correct_file is None or incorrect_file is None)):

        with st.spinner("Reading and comparing files…"):
            try:
                df_correct   = pd.read_excel(correct_file,   engine="openpyxl")
                df_incorrect = pd.read_excel(incorrect_file, engine="openpyxl")
            except Exception as e:
                st.error(f"Could not read Excel file(s): {e}")
                return

            # Align
            try:
                correct_aligned, incorrect_aligned, warnings = align_dataframes(
                    df_correct, df_incorrect, key_col
                )
            except ValueError as e:
                st.error(str(e))
                return

            for w in warnings:
                st.warning(w)

            # Compare
            match_mask = compare_dataframes(correct_aligned, incorrect_aligned, case_sensitive)

            # Summary stats
            total_cells    = match_mask.size
            total_matches  = int(match_mask.values.sum())
            total_mismatch = total_cells - total_matches
            match_pct      = 100 * total_matches / total_cells if total_cells else 0

            # ── Summary ───────────────────────────────────────────────────────
            st.subheader("📈 Summary")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Cells Compared", f"{total_cells:,}")
            m2.metric("✅ Matches",   f"{total_matches:,}",  delta=None)
            m3.metric("⚠️ Mismatches", f"{total_mismatch:,}", delta=None)
            m4.metric("Match Rate",  f"{match_pct:.1f}%")

            # ── Mismatch preview ─────────────────────────────────────────────
            if show_preview:
                diff_df = get_diff_preview(correct_aligned, incorrect_aligned, match_mask)
                if diff_df.empty:
                    st.success("🎉 No mismatches found — files are identical in the compared area!")
                else:
                    st.subheader(f"🔎 Mismatch Preview (up to 100 rows)")
                    st.dataframe(
                        diff_df.style.applymap(
                            lambda _: "background-color: #FFEB9C",
                            subset=["Incorrect Value"]
                        ).applymap(
                            lambda _: "background-color: #C6EFCE",
                            subset=["Correct Value"]
                        ),
                        use_container_width=True,
                        height=400,
                    )

            # ── Build and offer download ──────────────────────────────────────
            with st.spinner("Building formatted Excel report…"):
                xlsx_bytes = build_output_excel(incorrect_aligned, match_mask)

            st.subheader("⬇️ Download Report")
            st.download_button(
                label="Download Comparison Report (.xlsx)",
                data=xlsx_bytes,
                file_name="comparison_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success("Report ready! Click the button above to download.")

    elif correct_file is None or incorrect_file is None:
        st.info("Upload both files above to enable comparison.")


if __name__ == "__main__":
    main()
